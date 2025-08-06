import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# 📁 Detect Desktop path
def get_desktop_path():
    user_profile = os.environ.get("USERPROFILE") or os.path.expanduser("~")
    desktop_path = os.path.join(user_profile, "Desktop")
    return desktop_path if os.path.exists(desktop_path) else os.getcwd()

# 📅 Validate and normalize month input
def get_valid_month():
    valid_months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    while True:
        month_input = input("📅 Enter month name (e.g., August): ").strip().capitalize()
        if month_input in valid_months:
            return month_input
        else:
            print("⚠️ Invalid month. Please enter a valid month name (e.g., April, August).")

# 📊 Generate analytics from raw data
def generate_analytics(df):
    df['💰 Amount'] = pd.to_numeric(df['💰 Amount'], errors='coerce')
    df = df.dropna(subset=['💰 Amount'])

    analytics = df.groupby('📦 Category')['💰 Amount'].agg(['sum', 'min', 'max']).reset_index()
    analytics.columns = ['📦 Category', '💰 Total', '🔻 Min', '🔺 Max']

    analytics['📊 Previous Month'] = 3000
    analytics['📈 Change'] = analytics['💰 Total'] - analytics['📊 Previous Month']
    analytics['📌 Comparison'] = analytics['📈 Change'].apply(
        lambda x: "🔺 More than last month" if x > 0 else ("🔻 Less than last month" if x < 0 else "⚖️ Same as last month")
    )

    # ✅ Add total expense row with actual value in 💰 Total column
    total_expense = analytics['💰 Total'].sum()
    total_row = pd.DataFrame([{
        '📦 Category': '🧾 Total Expense This Month',
        '💰 Total': total_expense,
        '🔻 Min': '',
        '🔺 Max': '',
        '📊 Previous Month': '',
        '📈 Change': '',
        '📌 Comparison': ''
    }])

    return pd.concat([analytics, total_row], ignore_index=True)

# 🎨 Style Analytics sheet
def style_analytics_sheet(filepath):
    try:
        wb = load_workbook(filepath)
        ws = wb["Analytics"]

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        bold_font = Font(bold=True)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            comparison_cell = row[6]  # 📌 Comparison
            category_cell = row[0]

            if category_cell.value == "🧾 Total Expense This Month":
                try:
                    ws.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=3)
                except Exception as merge_error:
                    print(f"⚠️ Merge failed: {merge_error}")
                for cell in row:
                    cell.font = bold_font
                    cell.fill = gray_fill
            else:
                if comparison_cell.value.startswith("🔺"):
                    comparison_cell.fill = red_fill
                elif comparison_cell.value.startswith("🔻"):
                    comparison_cell.fill = green_fill
                elif comparison_cell.value.startswith("⚖️"):
                    comparison_cell.fill = gray_fill

        wb.save(filepath)
    except Exception as e:
        print(f"⚠️ Styling failed: {e}")

# 📐 Auto-adjust column widths
def auto_adjust_column_width(filepath):
    try:
        wb = load_workbook(filepath)
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        wb.save(filepath)
    except Exception as e:
        print(f"⚠️ Column width adjustment failed: {e}")

# 💾 Save to Excel
def save_to_excel(df, month):
    desktop = get_desktop_path()
    year = datetime.now().year
    filename = f"📊 Expenses_{month}_{year}.xlsx"
    filepath = os.path.join(desktop, filename)

    analytics_df = generate_analytics(df)

    try:
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name="📝 Raw Entries", index=False)
            analytics_df.to_excel(writer, sheet_name="Analytics", index=False)

        auto_adjust_column_width(filepath)
        style_analytics_sheet(filepath)

        wb = load_workbook(filepath)
        print(f"\n✅ Saved successfully to: {filepath}")
        print(f"📄 Sheets in file: {wb.sheetnames}")

    except Exception as e:
        print(f"❌ Error saving file: {e}")

# 🧑‍💻 Input loop
def get_user_input(month):
    entries = []
    print("\n📥 Enter your expenses (type 'done' to finish):")
    while True:
        category = input("📦 Category: ").strip()
        if category.lower() == 'done': break
        amount = input("💰 Amount: ").strip()
        notes = input("🗒️ Notes: ").strip()

        if not category or not amount:
            print("⚠️ Category and Amount are required. Try again.\n")
            continue

        try:
            amount = float(amount)
        except ValueError:
            print("⚠️ Amount must be a number. Try again.\n")
            continue

        today = datetime.now().strftime("%Y-%m-%d")
        entries.append({
            '📅 Date': today,
            '📆 Month': month,
            '📦 Category': category,
            '💰 Amount': amount,
            '🗒️ Notes': notes
        })
        print("✅ Entry added.\n")

    return pd.DataFrame(entries)

# 🚀 Main
if __name__ == "__main__":
    try:
        month = get_valid_month()
        df = get_user_input(month)
        if not df.empty:
            save_to_excel(df, month)
        else:
            print("⚠️ No data entered. Exiting.")
    except Exception as e:
        import traceback
        print("❌ Exception occurred:")
        traceback.print_exc()